import os
import json
import shutil
import subprocess
import base64
import re
import sys
import logging
from pathlib import Path

import numpy as np

LOG_LEVEL = os.getenv("VFP_LOG_LEVEL", "INFO").upper()
logging.basicConfig(
    level=getattr(logging, LOG_LEVEL, logging.INFO),
    format="%(asctime)s %(levelname)s [%(name)s] %(message)s",
    stream=sys.stdout,
    force=True,
)
logger = logging.getLogger("vfp-engine")

# Local module imports
from vfp_processing.downwashLLT import compute_downwash_LLT
from vfp_processing.readVFP import readFLOW, readFORCE, readCP, readWAVEDRG
from utils.VFP_File_Generation_Utils import extract_levels_and_fuse
from utils import VFP_Data_Extraction_Utils as vfp_extract

# --- CONFIGURATION ---
# vfpFilePath = r"C:\Users\Tarun.Ramprakash\Downloads\CRM-Wing-3.vfp"

if len(sys.argv) > 1:
    vfpFilePath = sys.argv[1]
    logger.debug("Using VFP file path from argv: %s", vfpFilePath)
else:
    raise RuntimeError("Usage: python vfp-engine.py <vfpFilePath>")

# --- UTILITY FUNCTIONS ---

def write_vfp_input_files(vfp_data, config_key, base_dir):
    logger.debug("Writing VFP input files for %s into %s", config_key, base_dir)
    input_files = vfp_data.get("inputFiles", {})
    config = input_files.get(config_key, {})
    file_names = config.get("fileNames", {})
    file_data = config.get("fileData", {})
    for key, file_name in file_names.items():
        if file_name and file_name in file_data:
            data = file_data[file_name].rstrip('\r\n')
            file_path = os.path.join(base_dir, file_name)
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            with open(file_path, "w", encoding="utf-8", newline='') as f:
                f.write(data)
            logger.debug("Wrote input file: %s", file_path)
        else:
            logger.debug("Skipping missing file for key %s in %s", key, config_key)

def copy_simulation_files(src_dir, dest_dir):
    logger.debug("Copying simulation helper files from %s to %s", src_dir, dest_dir)
    if not os.path.exists(src_dir):
        logger.warning("Source directory %s does not exist; skipping copy.", src_dir)
        return
    os.makedirs(dest_dir, exist_ok=True)
    for file_name in os.listdir(src_dir):
        src_file = os.path.join(src_dir, file_name)
        dest_file = os.path.join(dest_dir, file_name)
        if os.path.isfile(src_file):
            shutil.copy2(src_file, dest_file)
            logger.debug("Copied helper file %s -> %s", src_file, dest_file)

def add_vfp_results_to_data(vfp_data, sim_dir, config_key):
    """
    Adds result files for the current simulation to vfp_data, avoiding duplicates and only storing .mapout once.
    Each simulation's results are keyed by [flowFileName][result_file_name].
    Also adds wavedrag files: {geoFileName}{flowFileName}+wavedrg73/74/75/76.DAT
    """
    import base64

    result_exts = [
        ".mapout", ".cp", ".flow", ".conv", ".forces", ".sum", ".vis", ".DAT",
        ".fort11", ".fort15", ".fort21", ".fort50", ".fort51", ".fort52", ".fort55"
    ]
    wavedrag_suffixes = [f"wavedrg{n}.DAT" for n in range(73, 77)]

    file_names = vfp_data["inputFiles"][config_key]["fileNames"]
    geo_file = os.path.splitext(file_names["GeoFile"])[0]
    flow_file = file_names["DatFile"]
    flow_file_no_ext = os.path.splitext(flow_file)[0]
    flow_key = flow_file_no_ext  # use flow key without extension for results indexing

    # Prepare results dict — guard against both missing key and explicit None value
    if not isinstance(vfp_data.get("results"), dict):
        vfp_data["results"] = {}
    if not isinstance(vfp_data["results"].get(config_key), dict):
        vfp_data["results"][config_key] = {}
    if not isinstance(vfp_data["results"][config_key].get(flow_key), dict):
        vfp_data["results"][config_key][flow_key] = {}

    results = vfp_data["results"][config_key][flow_key]

    # Track if .mapout has already been stored for this config
    mapout_written = False
    for sim_results in vfp_data["results"][config_key].values():
        if any(fname.lower().endswith('.mapout') for fname in sim_results):
            mapout_written = True
            break

    # Standard result files
    for ext in result_exts:
        if ext == ".mapout":
            fname = f"{geo_file}{ext}"
            fpath = os.path.join(sim_dir, fname)
            if os.path.isfile(fpath) and not mapout_written:
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        data = f.read()
                    results[fname] = {"data": data, "encoding": "utf-8"}
                except Exception:
                    with open(fpath, "rb") as f:
                        bdata = f.read()
                    b64data = base64.b64encode(bdata).decode("ascii")
                    results[fname] = {"data": b64data, "encoding": "base64"}
                mapout_written = True
            continue

        fname = f"{geo_file}{flow_file_no_ext}{ext}"
        fpath = os.path.join(sim_dir, fname)
        if os.path.isfile(fpath):
            if fname not in results:
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        data = f.read()
                    results[fname] = {"data": data, "encoding": "utf-8"}
                except Exception:
                    with open(fpath, "rb") as f:
                        bdata = f.read()
                    b64data = base64.b64encode(bdata).decode("ascii")
                    results[fname] = {"data": b64data, "encoding": "base64"}

    # Wavedrag result files
    for suffix in wavedrag_suffixes:
        fname = f"{geo_file}{flow_file_no_ext}{suffix}"
        fpath = os.path.join(sim_dir, fname)
        if os.path.isfile(fpath):
            if fname not in results:
                try:
                    with open(fpath, "r", encoding="utf-8") as f:
                        data = f.read()
                    results[fname] = {"data": data, "encoding": "utf-8"}
                    logger.debug("Added wavedrag file to results: %s", fname)
                except Exception:
                    with open(fpath, "rb") as f:
                        bdata = f.read()
                    b64data = base64.b64encode(bdata).decode("ascii")
                    results[fname] = {"data": b64data, "encoding": "base64"}
    logger.info(
        "Attached %d result files for config %s flow %s", len(results), config_key, flow_file
    )
    return vfp_data


def extract_polars(sim_dir):
    """Build polar arrays from .forces and wavedrag files in a simulation directory."""
    forces_paths = [
        os.path.join(sim_dir, fname)
        for fname in os.listdir(sim_dir)
        if fname.lower().endswith(".forces")
    ]
    wavedrag_paths = [
        os.path.join(sim_dir, fname)
        for fname in os.listdir(sim_dir)
        if "wavedrg73" in fname.lower()
    ]

    forces_by_aoa = {}
    for fpath in forces_paths:
        aoa = vfp_extract.extract_aoa_from_forces_filename(fpath)
        if aoa is not None:
            forces_by_aoa[aoa] = fpath

    wavedrag_by_aoa = {}
    for wpath in wavedrag_paths:
        aoa = vfp_extract.extract_aoa_from_filename(wpath)
        if aoa is not None:
            wavedrag_by_aoa[aoa] = wpath

    common_aoas = sorted(set(forces_by_aoa.keys()) & set(wavedrag_by_aoa.keys()))
    if not common_aoas:
        logger.info("No matching .forces and wavedrag files found for polars extraction in %s", sim_dir)
        return {}

    records = []
    for aoa in common_aoas:
        fpath = forces_by_aoa[aoa]
        wpath = wavedrag_by_aoa[aoa]
        record = vfp_extract.extract_data(fpath, wpath)
        records.append(record)

    # Sort by ALPHA to keep polars ordered
    def _alpha_or_inf(rec):
        val = rec.get("ALPHA")
        return val if val is not None else float("inf")

    records = sorted(records, key=_alpha_or_inf)

    polar_keys = [
        "ALPHA",
        "MACH NO",
        "CL",
        "Cdv",
        "Cdi",
        "CDtotVFP",
        "CDtotIBE",
        "CMTOT(VFP)",
        "CDW_Upper",
        "CDW_Lower",
        "CDW(tot)",
    ]

    polars = {key: [] for key in polar_keys}
    polars["Filename"] = []

    for rec in records:
        for key in polar_keys:
            polars[key].append(rec.get(key))
        polars["Filename"].append(rec.get("Filename"))

    logger.info("Extracted polars for %d points from %s", len(records), sim_dir)
    return polars


def run_vfp(sim_dir, vfp_data, config_key):
    logger.info("Starting VFP run for %s in %s", config_key, sim_dir)
    bat_path = os.path.join(sim_dir, "cmdvfp.bat")
    simName = vfp_data.get("formData", {}).get("simName", "unknownSim")
    files = vfp_data["inputFiles"][config_key]["fileNames"]
    geoName = os.path.splitext(files["GeoFile"])[0]
    mapName = os.path.splitext(files["MapFile"])[0]
    datName = os.path.splitext(files["DatFile"])[0]
    continuation = vfp_data["formData"].get("continuationRun", False)
    excrescence = vfp_data["formData"].get("excrescence", False)
    cont_flag = "y" if continuation else "n"
    excr_flag = "y" if excrescence else "n"
    args = [bat_path, mapName, geoName, datName, excr_flag, cont_flag, ""]
    logger.debug("Subprocess args: %s", args)
    process = subprocess.Popen(
        args,
        cwd=sim_dir,
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True
    )
    for line in process.stdout:
        sys.stdout.write(line)
        sys.stdout.flush()
        logger.debug("VFP stdout: %s", line.rstrip())
    process.wait()
    logger.info("VFP run completed for %s with exit code %s", config_key, process.returncode)
    add_vfp_results_to_data(vfp_data, sim_dir, config_key)
    logger.info("Simulation results added to VFP Data for %s", config_key)
    save_vfp_results(vfp_data, simName=simName, aoa=None, project_root=project_root)
    return vfp_data

def vfp_dumpfile_write(sim_dir, vfp_data, config_key, dumpFileKey):
    required_exts = [".fort11", ".fort15", ".fort21", ".fort50", ".fort51", ".fort52", ".fort55"]
    results_root = vfp_data.get("results")
    if results_root is None or config_key not in results_root:
        available = list(results_root.keys()) if results_root else []
        raise ValueError(f"No results found for {config_key} in vfp_data. Available result configs: {available}")

    results_dict = results_root[config_key]
    if not results_dict:
        raise ValueError(f"Results for {config_key} are empty; flow keys available: {list(results_root.keys())}")

    if dumpFileKey not in results_dict:
        raise FileNotFoundError(
            f"Flow key '{dumpFileKey}' not found in results for {config_key}. "
            f"Available flow keys: {list(results_dict.keys())}"
        )

    logger.info("Writing dump files using flow key %s", dumpFileKey)
    results = results_dict[dumpFileKey]
    geoName = os.path.splitext(vfp_data["inputFiles"][config_key]["fileNames"]["GeoFile"])[0]
    dumpFileName = geoName + dumpFileKey 

    for ext in required_exts:
        # Find the first file in this flow entry that ends with the required extension
        matching_names = [name for name in results if name.lower().endswith(ext)]
        if not matching_names:
            raise FileNotFoundError(
                f"Required dump file {dumpFileName} with extension {ext} not found under flow key '{dumpFileKey}'."
            )
        fname = matching_names[0]
        file_info = results[fname]
        data = file_info["data"]
        encoding = file_info.get("encoding", "utf-8")
        file_path = os.path.join(sim_dir, fname)
        if encoding == "base64":
            with open(file_path, "wb") as f:
                f.write(base64.b64decode(data))
            logger.debug("Wrote binary dump file: %s", file_path)
        else:
            with open(file_path, "w", encoding="utf-8", newline='') as f:
                f.write(data)
            logger.debug("Wrote dump file: %s", file_path)


def write_dump_files_from_split_json(project_root, upload_id, split_file_name, sim_dir):
    """Write required dump files from a split-json artifact and return the dump base stem."""
    split_path = Path(project_root) / "data" / "uploads" / upload_id / "split-json" / split_file_name
    if not split_path.exists():
        raise FileNotFoundError(f"Split JSON file not found: {split_path}")

    with open(split_path, "r", encoding="utf-8") as f:
        payload = json.load(f)

    required_exts = {".fort11", ".fort15", ".fort21", ".fort50", ".fort51", ".fort52", ".fort55"}
    dump_base = None

    for fname, meta in payload.items():
        dest_path = Path(sim_dir) / fname
        encoding = meta.get("encoding", "utf-8")
        data = meta.get("data", "")
        if encoding == "base64":
            dest_path.write_bytes(base64.b64decode(data))
        else:
            dest_path.write_text(data, encoding="utf-8")

        if dump_base is None and dest_path.suffix.lower() in required_exts:
            dump_base = dest_path.stem

    if dump_base is None:
        raise ValueError("No dump files (.fort11/15/21/50/51/52/55) found in split JSON")

    logger.info("Restored dump files from split JSON %s into %s", split_file_name, sim_dir)
    return dump_base

def write_dump_files_from_payload(cont_dump_data: dict, sim_dir: str) -> str:
    """Write fort dump files delivered inline in continuationDumpData.

    The frontend stores only the seven fort files (fort11/15/21/50/51/52/55)
    in IndexedDB and sends them as ``continuationDumpData.files`` – a mapping
    of ``{filename: {data: str, encoding: 'utf-8'|'base64'}}``.
    Returns the dump base stem (filename without extension).
    """
    files = cont_dump_data.get("files", {})
    if not files:
        raise ValueError("continuationDumpData.files is empty")
    required_exts = {".fort11", ".fort15", ".fort21", ".fort50", ".fort51", ".fort52", ".fort55"}
    dump_base = None
    for fname, meta in files.items():
        dest_path = Path(sim_dir) / fname
        encoding = meta.get("encoding", "utf-8")
        data = meta.get("data", "")
        if encoding == "base64":
            dest_path.write_bytes(base64.b64decode(data))
        else:
            dest_path.write_text(data, encoding="utf-8")
        logger.debug("Wrote dump file from payload: %s", dest_path)
        if dump_base is None and dest_path.suffix.lower() in required_exts:
            dump_base = dest_path.stem
    if dump_base is None:
        raise ValueError(
            f"No fort dump files found in continuationDumpData.files. Got: {list(files.keys())}"
        )
    logger.info("Wrote %d dump files from payload, base: %s", len(files), dump_base)
    return dump_base


def run_vfp_continuation(sim_dir, vfp_data, config_key, dumpFileName):
    logger.info("Starting VFP continuation for %s using dump %s", config_key, dumpFileName)
    form_data = vfp_data.get("formData", {}) if isinstance(vfp_data, dict) else {}
    upload_id = form_data.get("uploadId")
    simName = vfp_data.get("formData", {}).get("simName", "unknownSim")
    split_file = form_data.get("continuationSplitFile")
    cont_dump_data = form_data.get("continuationDumpData")

    if cont_dump_data and cont_dump_data.get("configKey") == config_key:
        # Case 4 (new): dump files delivered inline from browser IndexedDB
        dump_base = write_dump_files_from_payload(cont_dump_data, sim_dir)
    elif upload_id and split_file:
        # Case 3: large file uploaded to server – dump files in split-json
        dump_base = write_dump_files_from_split_json(project_root, upload_id, split_file, sim_dir)
    else:
        # Case 2: small file – full results embedded in vfpData
        vfp_dumpfile_write(sim_dir, vfp_data, config_key, dumpFileName)
        dump_base = None
    bat_path = os.path.join(sim_dir, "cmdvfp.bat")
    files = vfp_data["inputFiles"][config_key]["fileNames"]
    geoName = os.path.splitext(files["GeoFile"])[0]
    mapName = os.path.splitext(files["MapFile"])[0]
    datName = os.path.splitext(files["DatFile"])[0]
    continuation = vfp_data["formData"].get("continuationRun", False)
    excrescence = vfp_data["formData"].get("excrescence", False)
    cont_flag = "y" if continuation else "n"
    excr_flag = "y" if excrescence else "n"
    if dump_base is None:
        dump_base = geoName + os.path.splitext(dumpFileName)[0]
    args = [bat_path, mapName, geoName, datName, excr_flag, cont_flag, dump_base, ""]
    logger.debug("Continuation subprocess args: %s", args)
    process = subprocess.Popen(
        args,
        cwd=sim_dir,
        shell=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True
    )
    for line in process.stdout:
        sys.stdout.write(line)
        sys.stdout.flush()
        logger.debug("VFP continuation stdout: %s", line.rstrip())
    process.wait()
    logger.info(
        "VFP continuation completed for %s with exit code %s", config_key, process.returncode
    )
    add_vfp_results_to_data(vfp_data, sim_dir, config_key)
    logger.info("Continuation results added to VFP Data for %s", config_key)
    save_vfp_results(vfp_data, simName=simName, aoa=None, project_root=project_root)
    return vfp_data

def extract_alpha_mach_from_flow(flow_file, level_idx=1):
    result = readFLOW(flow_file)
    level_key = f"level{level_idx}"
    first_line = result['levels'][level_key][0]
    parts = first_line.split()
    mach = float(parts[3])
    alpha = float(parts[4])
    return alpha, mach

def modify_tail_flow_file_preserve_format(tail_flow, new_aoa, new_mach):
    logger.debug("Modifying tail flow file %s with AoA=%s Mach=%s", tail_flow, new_aoa, new_mach)
    with open(tail_flow, 'r') as f:
        lines = f.readlines()
    data_json = extract_levels_and_fuse(lines)
    new_mach_str = f"{float(new_mach):.4f}"
    new_aoa_str = f"{float(new_aoa):.4f}"
    for key in data_json:
        if key.startswith('level'):
            block = data_json[key]
            first_line = block[0]
            pattern = re.compile(
                r'(^\s*2\s+\S+\s+\d+\s+)'  # Prefix up to Mach
                r'([-+]?\d+\.\d+)(\s+)'    # Mach number
                r'([-+]?\d+\.\d+)(\s+)'    # AoA
            )
            def repl(m):
                return m.group(1) + new_mach_str + m.group(3) + new_aoa_str + m.group(5)
            block[0] = pattern.sub(repl, first_line, count=1)
            data_json[key] = block
    new_lines = []
    new_lines.append(lines[0])
    new_lines.append(lines[1])
    idx = 2
    if 'fuse' in data_json:
        fuse_lines = data_json['fuse']
        new_lines.extend(fuse_lines)
        idx += len(fuse_lines)
    level_idx = 1
    while idx < len(lines):
        line = lines[idx]
        if line.lstrip().startswith('2') and f'level{level_idx}' in data_json:
            new_lines.extend(data_json[f'level{level_idx}'])
            idx += len(data_json[f'level{level_idx}'])
            level_idx += 1
        else:
            new_lines.append(line)
            idx += 1
    with open(tail_flow, 'w') as f:
        f.writelines(new_lines)
    logger.info("Tail flow file updated with preserved format: %s", tail_flow)

def _json_safe_default(obj):
    if isinstance(obj, np.ndarray):
        return obj.tolist()
    if isinstance(obj, (np.floating, np.integer, np.bool_)):
        return obj.item()
    return str(obj)


def save_vfp_results(vfp_data, simName, aoa, project_root):
    if aoa is not None:
        aoa_str = f"{float(aoa):.2f}".replace('.', 'p').replace('-', 'm')
        result_filename = f"{simName}-a{aoa_str}.vfp"
    else:
        result_filename = f"{simName}_results.vfp"
    outputVfpPath = os.path.join(project_root, "data", "Simulations", result_filename)
    with open(outputVfpPath, "w", encoding="utf-8") as f:
        json.dump(vfp_data, f, indent=4, default=_json_safe_default)
    logger.info("Saved VFP results to %s", outputVfpPath)

def cleanup_sim_dir(sim_dir):
    try:
        shutil.rmtree(sim_dir)
        logger.info("Deleted simulation working directory: %s", sim_dir)
    except Exception as e:
        logger.warning("Failed to delete simulation working directory %s: %s", sim_dir, e)


def generate_flow_file(base_flow_path, new_flow_path, new_aoa=None, new_mach=None):
    """
    Generate a new flow file with:
    1. Title line (file name)
    2. '    n' where n is the number of fuse lines (or 0)
    3. Fuse lines (if any)
    4. Level 1 lines (with modified Mach/AoA and second field's first char, preserving format)
    5. End with '    0'
    """
    import re
    logger.debug(
        "Generating flow file from %s to %s with AoA=%s Mach=%s",
        base_flow_path,
        new_flow_path,
        new_aoa,
        new_mach,
    )
    result = readFLOW(base_flow_path)
    levels_dict = result.get('levels', {})
    level_numbers = [int(m.group(1)) for key in levels_dict for m in re.finditer(r'^level(\d+)$', key)]
    if not level_numbers:
        raise ValueError("No level blocks found in flow file.")
    highest_level_key = f"level{max(level_numbers)}"
    level1_lines = levels_dict[highest_level_key][:]
    fuse_lines = result.get('fuse', None)
    # Modify the first line of level 1
    first_line = level1_lines[0]
    parts = first_line.split()
    if len(parts) < 5:
        raise ValueError("Unexpected format in flow file's level 1 first line.")
    # Modify Mach and AoA if needed
    if new_mach is not None:
        parts[3] = f"{float(new_mach):.4f}"
    if new_aoa is not None:
        parts[4] = f"{float(new_aoa):.4f}"
    # Change the first character of the second part to '1'
    if len(parts[1]) > 0:
        parts[1] = '1' + parts[1][1:]
    # Rebuild the first line, preserving original spacing
    def replace_nth_field(line, n, new_value):
        fields = list(re.finditer(r'\S+', line))
        if len(fields) > n:
            start, end = fields[n].span()
            return line[:start] + new_value + line[end:]
        return line
    mod_line = first_line
    mod_line = replace_nth_field(mod_line, 3, parts[3])
    mod_line = replace_nth_field(mod_line, 4, parts[4])
    # Replace the second field's first character
    fields = list(re.finditer(r'\S+', mod_line))
    if len(fields) > 1:
        start, end = fields[1].span()
        orig = mod_line[start:end]
        mod = '1' + orig[1:]
        mod_line = mod_line[:start] + mod + mod_line[end:]
    level1_lines[0] = mod_line

    with open(new_flow_path, "w", newline='') as f:
        # 1. Title line (file name)
        f.write("Flow File" + "\r\n")
        # 2. '    n' where n is the number of fuse lines (or 0)
        n_fuse = len(fuse_lines) if fuse_lines else 0
        f.write(f"   {n_fuse}\r\n")
        # 3. Fuse lines (if any)
        if fuse_lines and n_fuse > 0:
            for fuse_line in fuse_lines:
                f.write(fuse_line+"\n")
        # 4. Level 1 lines
        for line in level1_lines:
            f.write(line+"\n")
        # 5. End with '    0'
        f.write("    0\r\n")
    logger.info("Generated flow file (level 1 only) at %s", new_flow_path)


def get_flow_filename(mach, aoa):
    mach_str = f"{mach:.2f}".replace('.', '')
    aoa_sign = '+' if aoa >= 0 else '-'
    aoa_str = f"{abs(aoa):.2f}".replace('.', 'p')
    return f"M{mach_str}ma{aoa_sign}{aoa_str}.DAT"


def wing_step_succeeded(sim_dir, vfp_data, config_key):
    """
    Return True only when the wing .cp file exists **and** its level1 block
    contains at least one section with actual CP data.

    A crashed or unconverged run may still write a .cp file, but level1 will
    be absent or its sections will have empty CP lists.
    """
    try:
        files     = vfp_data["inputFiles"][config_key]["fileNames"]
        geo_stem  = os.path.splitext(files["GeoFile"])[0]
        flow_stem = os.path.splitext(files["DatFile"])[0]
        cp_path   = os.path.join(sim_dir, f"{geo_stem}{flow_stem}.cp")

        if not os.path.isfile(cp_path):
            logger.warning("Wing step: .cp file not found: %s", cp_path)
            return False

        cp_data = readCP(cp_path)
        level1  = cp_data.get("levels", {}).get("level1")
        if level1 is None:
            logger.warning("Wing step: level1 absent in .cp file: %s", cp_path)
            return False

        sections = level1.get("sections", {})
        # Check that at least one section has non-empty CP data
        for sec in sections.values():
            if sec.get("CP"):
                return True

        logger.warning(
            "Wing step: level1 present but no CP data in any section: %s", cp_path
        )
        return False

    except Exception as e:
        logger.warning("wing_step_succeeded check failed: %s", e)
        return False


def extract_forces_level1(forces_path):
    """
    Read a .forces file and return the Level-1 coefficient dict:
      {'alpha': float, 'mach': float, 'CL': float, 'CD_VFP': float,
       'CD_IBE': float, 'CM_VFP': float, 'Cdv': float, 'Cdi': float}
    Returns None if the file cannot be parsed.
    """
    try:
        data = readFORCE(forces_path)
        lv = data.get("levels", {}).get("level1")
        if lv is None:
            logger.warning("Level 1 not found in forces file: %s", forces_path)
            return None
        vfp  = lv.get("vfpCoefficients") or {}
        ibe  = lv.get("ibeCoefficients") or {}
        vrtx = lv.get("vortexCoefficients") or {}
        visc = lv.get("viscousDragData") or {}
        return {
            "alpha":    lv.get("alpha"),
            "mach":     lv.get("machNumber"),
            "CL":       ibe.get("CL") or vfp.get("CL"),
            "CD_VFP":   vfp.get("CD"),
            "CD_IBE":   ibe.get("CD"),
            "CM_VFP":   vfp.get("CM"),
            "Cdv":      visc.get("totalViscousDrag"),
            "Cdi":      vrtx.get("CD"),
        }
    except Exception as e:
        logger.warning("extract_forces_level1 failed for %s: %s", forces_path, e)
        return None


def extract_wavedrag_cdw(wavedrag_path):
    """
    Read a wavedrg73 file and return the combined CDW total:
      upper.cdwTotal + lower.cdwTotal  (both values already scaled by 1e-4)
    Returns None if the file is missing or cannot be parsed.
    """
    try:
        if not os.path.isfile(wavedrag_path):
            return None
        wd = readWAVEDRG(wavedrag_path)
        upper_cdw = wd.get("upperSurface", {}).get("cdwTotal")
        lower_cdw = wd.get("lowerSurface", {}).get("cdwTotal")
        if upper_cdw is None and lower_cdw is None:
            return None
        return (upper_cdw or 0.0) + (lower_cdw or 0.0)
    except Exception as e:
        logger.warning("extract_wavedrag_cdw failed for %s: %s", wavedrag_path, e)
        return None


def export_wing_tail_excel(rows, sim_name, project_root):
    """
    Write the combined wing+tail polar table to a formatted Excel file,
    using the same visual style as forces_to_excel.py.
    ``rows`` is a list of dicts, one per auto-runner step.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not rows:
            logger.warning("export_wing_tail_excel: no rows to write")
            return

        # ── Colour palette (matches forces_to_excel.py) ─────────────────
        CLR_HEADER_DARK  = "FF1F3864"
        CLR_HEADER_MED   = "FF2E75B6"
        CLR_HEADER_LIGHT = "FFD6E4F0"
        CLR_SCALAR_LABEL = "FF2F5496"
        CLR_SCALAR_BG    = "FFDCE6F1"
        CLR_ACCENT       = "FFED7D31"
        CLR_WHITE        = "FFFFFFFF"
        CLR_GREY_ROW     = "FFF2F2F2"
        CLR_WING_GRP     = "FF375623"   # forest green – wing group header
        CLR_TAIL_GRP     = "FF7030A0"   # purple – tail group header
        CLR_TOTAL_GRP    = "FF833C00"   # dark orange – totals group header

        def _thin_border():
            t = Side(style="thin")
            return Border(left=t, right=t, top=t, bottom=t)

        def _fill(hex_color):
            return PatternFill("solid", fgColor=hex_color)

        def _hdr_font(bold=True, color=CLR_WHITE, size=10):
            return Font(name="Calibri", bold=bold, color=color, size=size)

        def _body_font(bold=False, color="FF000000", size=10):
            return Font(name="Calibri", bold=bold, color=color, size=size)

        def _center():
            return Alignment(horizontal="center", vertical="center")

        def _right():
            return Alignment(horizontal="right", vertical="center")

        def _left():
            return Alignment(horizontal="left", vertical="center")

        def _write_merged_header(ws, row, col, title, ncols, fill_color):
            ws.merge_cells(start_row=row, start_column=col,
                           end_row=row, end_column=col + ncols - 1)
            c = ws.cell(row=row, column=col, value=title)
            c.font = _hdr_font(bold=True, size=10)
            c.fill = _fill(fill_color)
            c.alignment = _center()
            c.border = _thin_border()

        def _write_col_header(ws, row, col, label, fill_color):
            c = ws.cell(row=row, column=col, value=label)
            c.font = _hdr_font(color=CLR_WHITE, size=9)
            c.fill = _fill(fill_color)
            c.alignment = _center()
            c.border = _thin_border()

        def _write_data_cell(ws, row, col, value, alt=False):
            bg = CLR_GREY_ROW if alt else CLR_WHITE
            c = ws.cell(row=row, column=col, value=value)
            c.font = _body_font()
            c.fill = _fill(bg)
            c.border = _thin_border()
            if isinstance(value, float):
                c.number_format = "0.000000"
                c.alignment = _right()
            elif isinstance(value, int):
                c.alignment = _right()
            else:
                c.alignment = _left()

        # ── Column definitions ───────────────────────────────────────────
        # Each entry: (group_label, group_color, col_label, dict_key)
        COLS = [
            # Step / identity
            (None,          CLR_HEADER_DARK,  "Step",          "step"),
            # Wing aerodynamics
            (None,          CLR_WING_GRP,     "Wing AoA (°)",  "wing_aoa"),
            (None,          CLR_WING_GRP,     "Wing Mach",     "wing_mach"),
            ("WING",        CLR_WING_GRP,     "CL",            "wing_CL"),
            ("WING",        CLR_WING_GRP,     "CD (VFP)",      "wing_CD_VFP"),
            ("WING",        CLR_WING_GRP,     "CD (IBE)",      "wing_CD_IBE"),
            ("WING",        CLR_WING_GRP,     "CM (VFP)",      "wing_CM_VFP"),
            ("WING",        CLR_WING_GRP,     "Cd_visc",       "wing_Cdv"),
            ("WING",        CLR_WING_GRP,     "Cd_i (vortex)", "wing_Cdi"),
            ("WING",        CLR_WING_GRP,     "CDW (wave)",    "wing_CDW"),
            ("WING",        CLR_WING_GRP,     "CD_total",      "wing_CD_total"),
            # Downwash / tail identity
            (None,          CLR_TAIL_GRP,     "ε (°)",         "epsilon_deg"),
            (None,          CLR_TAIL_GRP,     "Tail AoA (°)",  "tail_alphat"),
            (None,          CLR_TAIL_GRP,     "Tail Mach",     "tail_mach"),
            ("TAIL",        CLR_TAIL_GRP,     "CL",            "tail_CL"),
            ("TAIL",        CLR_TAIL_GRP,     "CD (VFP)",      "tail_CD_VFP"),
            ("TAIL",        CLR_TAIL_GRP,     "CD (IBE)",      "tail_CD_IBE"),
            ("TAIL",        CLR_TAIL_GRP,     "CM (VFP)",      "tail_CM_VFP"),
            ("TAIL",        CLR_TAIL_GRP,     "Cd_visc",       "tail_Cdv"),
            ("TAIL",        CLR_TAIL_GRP,     "CDW (wave)",    "tail_CDW"),
            ("TAIL",        CLR_TAIL_GRP,     "CD_total",      "tail_CD_total"),
            # Aircraft totals
            ("TOTAL",       CLR_TOTAL_GRP,    "CL",            "total_CL"),
            ("TOTAL",       CLR_TOTAL_GRP,    "CD",            "total_CD"),
            ("TOTAL",       CLR_TOTAL_GRP,    "CM (VFP)",      "total_CM_VFP"),
        ]

        # Only include tail/total columns if there is any tail data in the rows
        has_tail_data = any(r.get("tail_alphat") is not None for r in rows)
        if not has_tail_data:
            COLS = [c for c in COLS if not (c[0] in ("TAIL", "TOTAL")
                                             or c[3] in ("epsilon_deg", "tail_alphat", "tail_mach"))]

        ncols = len(COLS)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AutoRunner Polar"
        ws.sheet_view.showGridLines = True

        # ── Row 1: Main title ────────────────────────────────────────────
        ws.row_dimensions[1].height = 24
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=ncols)
        title_cell = ws.cell(row=1, column=1,
                             value=f"VFP AutoRunner Results – {sim_name}")
        title_cell.font = _hdr_font(bold=True, color=CLR_WHITE, size=13)
        title_cell.fill = _fill(CLR_HEADER_DARK)
        title_cell.alignment = _center()
        title_cell.border = _thin_border()

        # ── Row 2: Group labels ──────────────────────────────────────────
        ws.row_dimensions[2].height = 16
        group_runs = []   # (start_col, end_col, label, color)
        run_start = 1
        run_label, run_color = COLS[0][0], COLS[0][1]
        for j, (grp, clr, _, _k) in enumerate(COLS[1:], start=2):
            if grp != run_label or clr != run_color:
                group_runs.append((run_start, j - 1, run_label, run_color))
                run_start = j
                run_label = grp
                run_color = clr
        group_runs.append((run_start, ncols, run_label, run_color))

        for (cs, ce, lbl, clr) in group_runs:
            if lbl:
                _write_merged_header(ws, 2, cs, lbl, ce - cs + 1, clr)
            else:
                for col in range(cs, ce + 1):
                    c = ws.cell(row=2, column=col, value="")
                    c.fill = _fill(clr)
                    c.border = _thin_border()

        # ── Row 3: Column headers ────────────────────────────────────────
        ws.row_dimensions[3].height = 18
        ws.freeze_panes = "A4"
        for j, (grp, clr, lbl, _k) in enumerate(COLS, start=1):
            _write_col_header(ws, 3, j, lbl, clr)

        # Auto-filter on header row
        ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"

        # ── Data rows ────────────────────────────────────────────────────
        for idx, row in enumerate(rows):
            r = 4 + idx
            ws.row_dimensions[r].height = 15
            alt = (idx % 2 == 1)
            for j, (_grp, _clr, _lbl, key) in enumerate(COLS, start=1):
                _write_data_cell(ws, r, j, row.get(key), alt=alt)

        # ── Column widths ────────────────────────────────────────────────
        KEY_WIDTHS = {
            "step": 6, "wing_aoa": 11, "wing_mach": 10,
            "wing_CL": 12, "wing_CD_VFP": 12, "wing_CD_IBE": 12,
            "wing_CM_VFP": 12, "wing_Cdv": 12, "wing_Cdi": 13,
            "wing_CDW": 12, "wing_CD_total": 13,
            "epsilon_deg": 10, "tail_alphat": 12, "tail_mach": 10,
            "tail_CL": 12, "tail_CD_VFP": 12, "tail_CD_IBE": 12,
            "tail_CM_VFP": 12, "tail_Cdv": 12, "tail_CDW": 12,
            "tail_CD_total": 13,
            "total_CL": 12, "total_CD": 12, "total_CM_VFP": 13,
        }
        for j, (_grp, _clr, _lbl, key) in enumerate(COLS, start=1):
            ws.column_dimensions[get_column_letter(j)].width = KEY_WIDTHS.get(key, 12)

        out_dir  = os.path.join(project_root, "data", "Simulations")
        out_path = os.path.join(out_dir, f"{sim_name}_AutoRunner_Forces.xlsx")
        wb.save(out_path)
        logger.info("Exported wing+tail combined forces to %s", out_path)

    except Exception as e:
        logger.error("export_wing_tail_excel failed: %s", e, exc_info=True)


# --- MAIN WORKFLOW ---

with open(vfpFilePath, "r", encoding="utf-8") as f:
    vfpData = json.load(f)

simName = vfpData["formData"]["simName"]
logger.info("Loaded VFP case: %s", simName)

project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
simBaseDir = os.path.join(project_root, "data", "Simulations", simName)
wingSimDir = os.path.join(simBaseDir, "wing")
tailSimDir = os.path.join(simBaseDir, "tail")
os.makedirs(wingSimDir, exist_ok=True)
os.makedirs(tailSimDir, exist_ok=True)

vfp_dir = os.path.join(project_root, "modules", "utils")
copy_simulation_files(vfp_dir, wingSimDir)
copy_simulation_files(vfp_dir, tailSimDir)
logger.info("Simulation helper files copied to working directories")

if "wingConfig" in vfpData.get("inputFiles", {}):
    write_vfp_input_files(vfpData, "wingConfig", base_dir=wingSimDir)
if "tailConfig" in vfpData.get("inputFiles", {}):
    write_vfp_input_files(vfpData, "tailConfig", base_dir=tailSimDir)
logger.info("All input files written to wing and tail working directories")

aoa = vfpData["formData"].get("aoa", None)
continuation = vfpData["formData"].get("continuationRun", False)
auto_runner = vfpData["formData"].get("autoRunner", False)

# --- Write tailSpec file if present in bodyFiles ---
tail_spec = None
if "bodyFiles" in vfpData["inputFiles"]:
    body_files = vfpData["inputFiles"]["bodyFiles"]
    if "fileNames" in body_files and "fileData" in body_files:
        for fname, fdata in body_files["fileData"].items():
            tail_spec_path = os.path.join(tailSimDir, fname)
            with open(tail_spec_path, "w", encoding="utf-8") as f:
                f.write(fdata)
            logger.debug("Wrote tail spec file: %s", tail_spec_path)
            tail_spec = tail_spec_path  # Use the first one found

# --- MODE SELECTION AND EXECUTION ---
if auto_runner:
    logger.info("AutoRunner mode enabled")
    autoMode     = vfpData["formData"].get("autoMode", "aoa")  # "aoa" or "mach"
    autoStepSize = float(vfpData["formData"].get("autoStepSize", 1.0))
    autoEndValue = (
        float(vfpData["formData"].get("autoEndAoA",  0.0)) if autoMode == "aoa"
        else float(vfpData["formData"].get("autoEndMach", 0.0))
    )
    base_aoa  = float(vfpData["formData"].get("aoa",  0.0))
    base_mach = float(vfpData["formData"].get("mach", 0.0))

    # Extract Reynolds number from the original flow file name
    orig_flow_file_name = vfpData["inputFiles"]["wingConfig"]["fileNames"]["DatFile"]
    re_match   = re.search(r"Re(\d+p\d+)", orig_flow_file_name)
    reynolds   = re_match.group(1) if re_match else "0p00"

    # Build the full sweep list
    n = int(round((
        (autoEndValue - base_aoa) if autoMode == "aoa" else (autoEndValue - base_mach)
    ) / autoStepSize)) + 1
    values = [
        base_aoa  + i * autoStepSize if autoMode == "aoa"
        else base_mach + i * autoStepSize
        for i in range(n)
    ]

    # --- Helper: build a standardised flow-file name for a given step ---
    def _std_flow_name(aoa_val, mach_val):
        mp  = f"M{mach_val:.2f}".replace(".", "")
        aps = f"{abs(aoa_val):05.2f}".replace(".", "p")
        sgn = "+" if aoa_val >= 0 else "-"
        return f"{mp}Re{reynolds}ma{sgn}{aps}.DAT"

    # Ensure step-0 flow file uses the standardised naming scheme.
    # If the user-supplied name already matches, nothing is copied.
    if autoMode == "aoa":
        step0_aoa, step0_mach = values[0], base_mach
    else:
        step0_aoa, step0_mach = base_aoa, values[0]
    std_name_0 = _std_flow_name(step0_aoa, step0_mach)
    if orig_flow_file_name != std_name_0:
        src0 = os.path.join(wingSimDir, orig_flow_file_name)
        dst0 = os.path.join(wingSimDir, std_name_0)
        if os.path.isfile(src0) and not os.path.isfile(dst0):
            shutil.copy2(src0, dst0)
            logger.info("Copied step-0 flow file %s -> %s", orig_flow_file_name, std_name_0)
    step0_flow_name = std_name_0
    base_flow_path  = os.path.join(wingSimDir, step0_flow_name)

    # --- Tail setup: check once whether a tail config is present ---
    has_tail = (
        "tailConfig" in vfpData.get("inputFiles", {})
        and vfpData["inputFiles"]["tailConfig"]["fileNames"].get("DatFile")
    )
    orig_tail_flow_name    = None
    orig_tail_flow_stem    = None
    orig_tail_flow_content = None
    tail_geo_path          = None
    tail_geo_stem          = None
    if has_tail:
        orig_tail_flow_name = vfpData["inputFiles"]["tailConfig"]["fileNames"]["DatFile"]
        orig_tail_flow_stem = os.path.splitext(orig_tail_flow_name)[0]
        orig_tail_flow_path = os.path.join(tailSimDir, orig_tail_flow_name)
        try:
            with open(orig_tail_flow_path, "r", encoding="utf-8") as _f:
                orig_tail_flow_content = _f.read()
        except Exception as _e:
            logger.error("AutoRunner: cannot read tail flow template %s: %s", orig_tail_flow_path, _e)
            has_tail = False
        tail_geo_file = vfpData["inputFiles"]["tailConfig"]["fileNames"]["GeoFile"]
        tail_geo_path = os.path.join(tailSimDir, tail_geo_file)
        tail_geo_stem = os.path.splitext(tail_geo_file)[0]
        logger.info(
            "AutoRunner: tail config detected; template=%s geo=%s",
            orig_tail_flow_name, tail_geo_path,
        )
    else:
        logger.info("AutoRunner: no tail config; wing-only polar sweep")

    # Accumulated rows for the combined Excel file (wing + tail per step)
    excel_rows = []

    # --- Helper: run tail for one wing step, returns row dict or None ---
    def _run_tail_for_step(step_idx, wing_aoa_val, wing_flow_stem):
        """
        Compute downwash, build a unique tail flow file, run the tail simulation,
        and return a partial result dict for the Excel row.  Returns None on any failure.
        """
        wing_geo_no_ext = os.path.splitext(
            vfpData["inputFiles"]["wingConfig"]["fileNames"]["GeoFile"]
        )[0]
        cp_fname = wing_geo_no_ext + wing_flow_stem + ".cp"
        cp_path  = os.path.join(wingSimDir, cp_fname)
        if not os.path.isfile(cp_path):
            logger.warning(
                "AutoRunner tail step %d: .cp not found (%s) – skipping tail",
                step_idx, cp_path,
            )
            return None
        try:
            logger.info(
                "AutoRunner tail step %d: downwash for wing_aoa=%.4f using %s",
                step_idx, wing_aoa_val, cp_path,
            )
            downwash_results = compute_downwash_LLT(
                cp_path, tail_geo_path, tail_spec, save_plots=False
            )
            ALPHAE_i = downwash_results.get("effective_epsilon_deg")
            MACHT_i  = round(downwash_results.get("avg_local_mach"), 4)
            ALPHAT_i = round(wing_aoa_val - abs(ALPHAE_i), 4)
            logger.info(
                "AutoRunner tail step %d: epsilon=%.4f alphat=%.4f mach_t=%.4f",
                step_idx, ALPHAE_i, ALPHAT_i, MACHT_i,
            )
        except Exception as _e:
            logger.error("AutoRunner tail step %d: downwash failed: %s", step_idx, _e)
            return None

        # Store downwash under results
        vfpData.setdefault("results", {}).setdefault("tailConfig", {}).setdefault(
            "downwash", {}
        )[f"wingAoA_{wing_aoa_val:.4f}"] = downwash_results

        # Build a unique, descriptive tail flow file name for this step:
        # e.g.  HtailM080ma-01p23.DAT  (includes tail mach and tail AoA)
        aoa_sign_t   = "+" if ALPHAT_i >= 0 else "-"
        alphat_str_t = f"{abs(ALPHAT_i):05.2f}".replace(".", "p")
        mach_part_t  = f"M{MACHT_i:.2f}".replace(".", "")
        tail_flow_name_i = (
            f"{orig_tail_flow_stem}_{mach_part_t}ma{aoa_sign_t}{alphat_str_t}.DAT"
        )
        tail_flow_path_i = os.path.join(tailSimDir, tail_flow_name_i)

        try:
            with open(tail_flow_path_i, "w", encoding="utf-8") as _tf:
                _tf.write(orig_tail_flow_content)
            modify_tail_flow_file_preserve_format(tail_flow_path_i, ALPHAT_i, MACHT_i)
        except Exception as _e:
            logger.error(
                "AutoRunner tail step %d: could not write/modify tail flow file: %s", step_idx, _e
            )
            return None

        saved_continuation = vfpData["formData"].get("continuationRun", False)
        vfpData["formData"]["continuationRun"] = False
        vfpData["inputFiles"]["tailConfig"]["fileNames"]["DatFile"] = tail_flow_name_i
        try:
            logger.info(
                "AutoRunner tail step %d: running tail alphat=%.4f mach=%.4f",
                step_idx, ALPHAT_i, MACHT_i,
            )
            run_vfp(tailSimDir, vfpData, "tailConfig")
        except Exception as _e:
            logger.error("AutoRunner tail step %d: run_vfp failed: %s", step_idx, _e)
            vfpData["formData"]["continuationRun"] = saved_continuation
            return None
        finally:
            vfpData["formData"]["continuationRun"] = saved_continuation

        # Check tail forces and wave drag files were produced
        tail_flow_stem_i = os.path.splitext(tail_flow_name_i)[0]
        tail_forces_path = os.path.join(
            tailSimDir, f"{tail_geo_stem}{tail_flow_stem_i}.forces"
        )
        tail_forces = extract_forces_level1(tail_forces_path)
        if tail_forces is None:
            logger.warning(
                "AutoRunner tail step %d: forces file not found or unreadable: %s",
                step_idx, tail_forces_path,
            )

        tail_wavedrag_path = os.path.join(
            tailSimDir, f"{tail_geo_stem}{tail_flow_stem_i}wavedrg73.DAT"
        )
        tail_cdw = extract_wavedrag_cdw(tail_wavedrag_path)
        if tail_cdw is not None:
            logger.info(
                "AutoRunner tail step %d: CDW(wave)=%.6f", step_idx, tail_cdw
            )
        else:
            logger.debug(
                "AutoRunner tail step %d: no wavedrag73 file found: %s",
                step_idx, tail_wavedrag_path,
            )

        return {
            "alphat":         ALPHAT_i,
            "tail_mach":      MACHT_i,
            "epsilon_deg":    ALPHAE_i,
            "tail_forces":    tail_forces,
            "tail_cdw":       tail_cdw,
            "tail_flow_stem": tail_flow_stem_i,
        }

    # ------------------------------------------------------------------ #
    # STEP 0 – first simulation (non-continuation)                        #
    # ------------------------------------------------------------------ #
    logger.info("AutoRunner step 0/%d at %s=%.4f", n - 1, autoMode, values[0])
    vfpData["inputFiles"]["wingConfig"]["fileNames"]["DatFile"] = step0_flow_name
    try:
        vfpData = run_vfp(wingSimDir, vfpData, "wingConfig")
        vfpData = add_vfp_results_to_data(vfpData, wingSimDir, "wingConfig")
    except Exception as _e:
        logger.error("AutoRunner step 0: wing run failed: %s", _e)

    wing_ok_0 = wing_step_succeeded(wingSimDir, vfpData, "wingConfig")
    wing_geo_stem = os.path.splitext(
        vfpData["inputFiles"]["wingConfig"]["fileNames"]["GeoFile"]
    )[0]
    step0_flow_stem = os.path.splitext(step0_flow_name)[0]

    wing_forces_0 = None
    wing_cdw_0    = None
    tail_result_0 = None
    if wing_ok_0:
        wing_forces_0 = extract_forces_level1(
            os.path.join(wingSimDir, f"{wing_geo_stem}{step0_flow_stem}.forces")
        )
        wing_cdw_0 = extract_wavedrag_cdw(
            os.path.join(wingSimDir, f"{wing_geo_stem}{step0_flow_stem}wavedrg73.DAT")
        )
        if has_tail and orig_tail_flow_content:
            tail_result_0 = _run_tail_for_step(0, step0_aoa, step0_flow_stem)
    else:
        logger.warning("AutoRunner step 0: wing simulation did not produce output – skipping tail")

    # Build Excel row for step 0
    row0 = {"step": 0, "wing_aoa": step0_aoa, "wing_mach": step0_mach}
    if wing_forces_0:
        row0.update({f"wing_{k}": v for k, v in wing_forces_0.items() if k not in ("alpha", "mach")})
        row0["wing_aoa"]  = wing_forces_0.get("alpha", step0_aoa)
        row0["wing_mach"] = wing_forces_0.get("mach",  step0_mach)
    row0["wing_CDW"] = wing_cdw_0
    if wing_forces_0:
        row0["wing_CD_total"] = (
            (wing_forces_0.get("Cdi") or 0.0)
            + (wing_forces_0.get("Cdv") or 0.0)
            + (wing_cdw_0 or 0.0)
        )
    if tail_result_0:
        row0["tail_alphat"] = tail_result_0["alphat"]
        row0["tail_mach"]   = tail_result_0["tail_mach"]
        row0["epsilon_deg"] = tail_result_0["epsilon_deg"]
        tf0 = tail_result_0["tail_forces"]
        if tf0:
            row0.update({f"tail_{k}": v for k, v in tf0.items() if k not in ("alpha", "mach")})
        row0["tail_CDW"] = tail_result_0.get("tail_cdw")
        row0["tail_CD_total"] = (
            ((tf0.get("Cdi") or 0.0) + (tf0.get("Cdv") or 0.0) + (tail_result_0.get("tail_cdw") or 0.0))
            if tf0 else None
        )
    # Combined totals
    if wing_forces_0:
        tf0 = tail_result_0.get("tail_forces") if tail_result_0 else None
        w_cl = wing_forces_0.get("CL") or 0.0
        t_cl = (tf0 or {}).get("CL") or 0.0
        row0["total_CL"] = (
            w_cl + t_cl if wing_forces_0.get("CL") is not None or (tf0 and tf0.get("CL") is not None) else None
        )
        row0["total_CM_VFP"] = (
            (wing_forces_0.get("CM_VFP") or 0.0) + ((tf0 or {}).get("CM_VFP") or 0.0)
        )
        row0["total_CD"] = (
            (row0.get("wing_CD_total") or 0.0) + (row0.get("tail_CD_total") or 0.0)
            if row0.get("wing_CD_total") is not None else None
        )
    excel_rows.append(row0)

    # ------------------------------------------------------------------ #
    # STEPS 1..n-1 – continuation runs                                    #
    # ------------------------------------------------------------------ #
    for i in range(1, n):
        if autoMode == "aoa":
            aoa_val  = values[i]
            mach_val = base_mach
        else:
            aoa_val  = base_aoa
            mach_val = values[i]

        flow_file_name = _std_flow_name(aoa_val, mach_val)
        flow_file_path = os.path.join(wingSimDir, flow_file_name)
        logger.info(
            "AutoRunner step %d/%d at %s=%.4f  flow=%s",
            i, n - 1, autoMode, values[i], flow_file_name,
        )

        # Generate the flow file for this step (continuation uses previous dump)
        try:
            generate_flow_file(
                base_flow_path, flow_file_path,
                new_aoa=aoa_val  if autoMode == "aoa"  else base_aoa,
                new_mach=mach_val if autoMode == "mach" else base_mach,
            )
        except Exception as _e:
            logger.error("AutoRunner step %d: generate_flow_file failed: %s", i, _e)
            excel_rows.append({"step": i, "wing_aoa": aoa_val, "wing_mach": mach_val,
                               "error": str(_e)})
            continue

        vfpData["inputFiles"]["wingConfig"]["fileNames"]["DatFile"] = flow_file_name

        # Determine which previous step's dump files to use
        if i == 1:
            dump_stem = os.path.splitext(step0_flow_name)[0]
        else:
            prev_name = _std_flow_name(
                values[i - 1] if autoMode == "aoa" else base_aoa,
                base_mach      if autoMode == "aoa" else values[i - 1],
            )
            dump_stem = os.path.splitext(prev_name)[0]

        vfpData["formData"]["continuationRun"] = True
        wing_ok_i = False
        try:
            vfpData = run_vfp_continuation(wingSimDir, vfpData, "wingConfig", dump_stem)
            vfpData = add_vfp_results_to_data(vfpData, wingSimDir, "wingConfig")
            wing_ok_i = wing_step_succeeded(wingSimDir, vfpData, "wingConfig")
        except Exception as _e:
            logger.error("AutoRunner step %d: wing continuation failed: %s", i, _e)
            excel_rows.append({"step": i, "wing_aoa": aoa_val, "wing_mach": mach_val,
                               "error": str(_e)})
            continue

        flow_stem_i   = os.path.splitext(flow_file_name)[0]
        wing_forces_i = None
        wing_cdw_i    = None
        tail_result_i = None
        if wing_ok_i:
            wing_forces_i = extract_forces_level1(
                os.path.join(wingSimDir, f"{wing_geo_stem}{flow_stem_i}.forces")
            )
            wing_cdw_i = extract_wavedrag_cdw(
                os.path.join(wingSimDir, f"{wing_geo_stem}{flow_stem_i}wavedrg73.DAT")
            )
            if has_tail and orig_tail_flow_content:
                try:
                    tail_result_i = _run_tail_for_step(i, aoa_val, flow_stem_i)
                except Exception as _e:
                    logger.error("AutoRunner step %d: tail step failed: %s", i, _e)
        else:
            logger.warning(
                "AutoRunner step %d: wing did not produce output – skipping tail", i
            )

        # Build Excel row
        row_i = {"step": i, "wing_aoa": aoa_val, "wing_mach": mach_val}
        if wing_forces_i:
            row_i.update({f"wing_{k}": v for k, v in wing_forces_i.items()
                          if k not in ("alpha", "mach")})
            row_i["wing_aoa"]  = wing_forces_i.get("alpha", aoa_val)
            row_i["wing_mach"] = wing_forces_i.get("mach",  mach_val)
        row_i["wing_CDW"] = wing_cdw_i
        if wing_forces_i:
            row_i["wing_CD_total"] = (
                (wing_forces_i.get("Cdi") or 0.0)
                + (wing_forces_i.get("Cdv") or 0.0)
                + (wing_cdw_i or 0.0)
            )
        if tail_result_i:
            row_i["tail_alphat"] = tail_result_i["alphat"]
            row_i["tail_mach"]   = tail_result_i["tail_mach"]
            row_i["epsilon_deg"] = tail_result_i["epsilon_deg"]
            tf_i = tail_result_i["tail_forces"]
            if tf_i:
                row_i.update({f"tail_{k}": v for k, v in tf_i.items()
                              if k not in ("alpha", "mach")})
            row_i["tail_CDW"] = tail_result_i.get("tail_cdw")
            row_i["tail_CD_total"] = (
                ((tf_i.get("Cdi") or 0.0) + (tf_i.get("Cdv") or 0.0) + (tail_result_i.get("tail_cdw") or 0.0))
                if tf_i else None
            )
        # Combined totals
        if wing_forces_i:
            tf_i = tail_result_i.get("tail_forces") if tail_result_i else None
            w_cl = wing_forces_i.get("CL") or 0.0
            t_cl = (tf_i or {}).get("CL") or 0.0
            row_i["total_CL"] = (
                w_cl + t_cl if wing_forces_i.get("CL") is not None or (tf_i and tf_i.get("CL") is not None) else None
            )
            row_i["total_CM_VFP"] = (
                (wing_forces_i.get("CM_VFP") or 0.0) + ((tf_i or {}).get("CM_VFP") or 0.0)
            )
            row_i["total_CD"] = (
                (row_i.get("wing_CD_total") or 0.0) + (row_i.get("tail_CD_total") or 0.0)
                if row_i.get("wing_CD_total") is not None else None
            )
        excel_rows.append(row_i)

    # ------------------------------------------------------------------ #
    # Post-sweep: polars, Excel export, and final .vfp save               #
    # ------------------------------------------------------------------ #
    try:
        polars_data = extract_polars(wingSimDir)
        if polars_data:
            vfpData.setdefault("results", {})["Polars"] = polars_data
        else:
            logger.warning("No polars extracted; results->Polars not updated")
    except Exception as _e:
        logger.error("AutoRunner: extract_polars failed: %s", _e)

    if excel_rows:
        export_wing_tail_excel(excel_rows, simName, project_root)

    try:
        save_vfp_results(vfpData, simName, values[-1], project_root)
    except Exception as _e:
        logger.error("AutoRunner: save_vfp_results failed: %s", _e)

    logger.info("AutoRunner complete – %d steps, %d Excel rows written", n, len(excel_rows))

elif continuation:
    logger.info("Continuation mode enabled")
    _upload_id   = vfpData["formData"].get("uploadId")
    _cont_dump   = vfpData["formData"].get("continuationDumpData")  # Case 4: inline payload
    if "wingConfig" in vfpData.get("inputFiles", {}):
        if _cont_dump and _cont_dump.get("configKey") == "wingConfig":
            # Case 4: dump files delivered inline from browser IndexedDB selection
            dumpFileName = _cont_dump.get("flowKey")
        elif _upload_id:
            # Case 3: large file uploaded to server – dump files are in split-json
            dumpFileName = vfpData["formData"].get("continuationSplitKey")
        else:
            # Case 2: small file parsed by client – full results embedded in vfpData
            dumpFileName = vfpData["formData"].get("wingDumpName", vfpData["formData"].get("dumpName"))
        logger.info("Simulating wing continuation with dump %s", dumpFileName)
        vfpData = run_vfp_continuation(wingSimDir, vfpData, "wingConfig", dumpFileName)
    if (
        "tailConfig" in vfpData.get("inputFiles", {}) and
        vfpData["inputFiles"]["tailConfig"]["fileNames"].get("DatFile")
    ):
        if _cont_dump and _cont_dump.get("configKey") == "tailConfig":
            # Case 4: dump files delivered inline from browser IndexedDB selection
            dumpFileName = _cont_dump.get("flowKey")
        elif _upload_id:
            # Case 3: large file uploaded to server – dump files are in split-json
            dumpFileName = vfpData["formData"].get("continuationSplitKey")
        else:
            # Case 2: small file parsed by client – full results embedded in vfpData
            dumpFileName = vfpData["formData"].get("tailDumpName", vfpData["formData"].get("dumpName"))
        vfpData = run_vfp_continuation(tailSimDir, vfpData, "tailConfig", dumpFileName)
    save_vfp_results(vfpData, simName, aoa, project_root)

else:
    logger.info("Standard mode enabled")
    if "wingConfig" not in vfpData.get("inputFiles", {}):
        raise RuntimeError("No wingConfig found in VFP data.")

    logger.info("Simulating wing (standard mode)")
    vfpData = run_vfp(wingSimDir, vfpData, "wingConfig")

    if (
        "tailConfig" in vfpData.get("inputFiles", {}) and
        vfpData["inputFiles"]["tailConfig"]["fileNames"].get("DatFile")
    ):
        logger.info("Computing downwash and modifying tail flow")

        wing_flow = os.path.join(wingSimDir, vfpData["inputFiles"]["wingConfig"]["fileNames"]["DatFile"])
        tail_flow = os.path.join(tailSimDir, vfpData["inputFiles"]["tailConfig"]["fileNames"]["DatFile"])
        tail_geo  = os.path.join(tailSimDir, vfpData["inputFiles"]["tailConfig"]["fileNames"]["GeoFile"])
        tail_map  = os.path.join(tailSimDir, vfpData["inputFiles"]["tailConfig"]["fileNames"]["MapFile"])

        ALPHAW, MACHW = extract_alpha_mach_from_flow(wing_flow)
        cp_files = [f for f in os.listdir(wingSimDir) if f.lower().endswith('.cp')]
        if not cp_files:
            raise RuntimeError("No .cp file found after wing simulation.")
        cp_file = os.path.join(wingSimDir, cp_files[0])

        downwash_results = compute_downwash_LLT(cp_file, tail_geo, tail_spec, save_plots=False)
        ALPHAE = downwash_results.get('effective_epsilon_deg')
        MACHT = downwash_results.get('avg_local_mach')
        ALPHAT = round(ALPHAW - abs(ALPHAE), 4)
        MACHT = round(MACHT, 4)
        logger.info(
            "Downwash results ALPHAW=%s ALPHAE=%s ALPHAT=%s MACHT=%s",
            ALPHAW,
            ALPHAE,
            ALPHAT,
            MACHT,
        )

        if "results" not in vfpData:
            vfpData["results"] = {}
        if "tailConfig" not in vfpData["results"]:
            vfpData["results"]["tailConfig"] = {}
        vfpData["results"]["tailConfig"]["flowLLT"] = downwash_results

        modify_tail_flow_file_preserve_format(tail_flow, ALPHAT, MACHT)

        with open(tail_flow, "r", encoding="utf-8") as f:
            tail_flow_data = f.read()
        vfpData["inputFiles"]["tailConfig"]["fileData"][os.path.basename(tail_flow)] = tail_flow_data

        logger.info("Simulating tail (standard mode)")
        vfpData = run_vfp(tailSimDir, vfpData, "tailConfig")

    save_vfp_results(vfpData, simName, aoa, project_root)

# --- CLEANUP (commented out for debugging) ---
# cleanup_sim_dir(simBaseDir)

