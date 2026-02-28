"""
post_export_excel.py
--------------------
Reconstruct the AutoRunner Excel polar export from an existing simulation
folder.  Use this when the simulation has already finished but the Excel
file was not generated (e.g. the Excel export feature was added after the
run completed).

Usage
-----
    python post_export_excel.py <simName>
    python post_export_excel.py <simName> --project-root <path>

    <simName>        Name of the simulation (folder under data/Simulations/).
    --project-root   Root of the VFP-Python repo. Defaults to the parent of
                     the directory containing this script.

The script looks for:
    data/Simulations/<simName>/wing/*.forces   (wing aerodynamic results)
    data/Simulations/<simName>/wing/*wavedrg73*  (wing wave-drag results)
    data/Simulations/<simName>/tail/*.forces   (tail results, optional)
    data/Simulations/<simName>/tail/*wavedrg73*  (tail wave-drag, optional)

Output
------
    data/Simulations/<simName>_AutoRunner_Forces.xlsx
"""

import argparse
import os
import re
import sys

# ---------------------------------------------------------------------------
# Resolve the project root so local modules can be imported
# ---------------------------------------------------------------------------
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
if _SCRIPT_DIR not in sys.path:
    sys.path.insert(0, _SCRIPT_DIR)

from vfp_processing.readVFP import readFORCE, readWAVEDRG


# ---------------------------------------------------------------------------
# Helpers (same logic as in vfp-engine.py)
# ---------------------------------------------------------------------------

def _extract_forces_level1(forces_path):
    """Return Level-1 coefficient dict from a .forces file, or None."""
    try:
        data = readFORCE(forces_path)
        lv = data.get("levels", {}).get("level1")
        if lv is None:
            return None
        vfp  = lv.get("vfpCoefficients") or {}
        ibe  = lv.get("ibeCoefficients")  or {}
        vrtx = lv.get("vortexCoefficients") or {}
        visc = lv.get("viscousDragData")  or {}
        return {
            "alpha":  lv.get("alpha"),
            "mach":   lv.get("machNumber"),
            "CL":     ibe.get("CL") or vfp.get("CL"),
            "CD_VFP": vfp.get("CD"),
            "CD_IBE": ibe.get("CD"),
            "CM_VFP": vfp.get("CM"),
            "Cdv":    visc.get("totalViscousDrag"),
            "Cdi":    vrtx.get("CD"),
        }
    except Exception as exc:
        print(f"  WARNING: could not read forces file {forces_path}: {exc}")
        return None


def _extract_wavedrag_cdw(wavedrag_path):
    """Return upper_cdwTotal + lower_cdwTotal from a wavedrg73 file, or None."""
    try:
        if not os.path.isfile(wavedrag_path):
            return None
        wd = readWAVEDRG(wavedrag_path)
        upper = wd.get("upperSurface", {}).get("cdwTotal")
        lower = wd.get("lowerSurface", {}).get("cdwTotal")
        if upper is None and lower is None:
            return None
        return (upper or 0.0) + (lower or 0.0)
    except Exception as exc:
        print(f"  WARNING: could not read wavedrag file {wavedrag_path}: {exc}")
        return None


def _find_partner_wavedrag(forces_path):
    """
    Given a .forces file path, return the path of the matching wavedrg73 file
    in the same directory (same stem prefix), or None.
    """
    directory = os.path.dirname(forces_path)
    stem = os.path.splitext(os.path.basename(forces_path))[0]
    for fn in os.listdir(directory):
        if "wavedrg73" in fn.lower() and fn.startswith(stem):
            return os.path.join(directory, fn)
    return None


def _scan_forces_files(sim_dir):
    """
    Return a list of (forces_path, wavedrag_path_or_None) tuples found in
    *sim_dir*, sorted by the alpha read from each forces file.
    Files whose level1 cannot be parsed are skipped with a warning.
    """
    entries = []
    for fn in os.listdir(sim_dir):
        if not fn.lower().endswith(".forces"):
            continue
        fp = os.path.join(sim_dir, fn)
        f = _extract_forces_level1(fp)
        if f is None:
            print(f"  WARNING: skipping unreadable forces file: {fn}")
            continue
        wd_path = _find_partner_wavedrag(fp)
        entries.append((f["alpha"], fp, wd_path, f))

    # Sort by alpha (ascending)
    entries.sort(key=lambda x: x[0] if x[0] is not None else float("inf"))
    return [(fp, wd, f) for (_alpha, fp, wd, f) in entries]


# ---------------------------------------------------------------------------
# Row builder
# ---------------------------------------------------------------------------

def build_rows(wing_entries, tail_entries):
    """
    Combine wing and tail entries into a list of row dicts.
    Matching is done by sort order (step index) — the AutoRunner processes
    both in the same ascending-AoA order.
    Tail entries list may be empty (wing-only case).
    """
    rows = []
    for step, (wing_fp, wing_wd_path, wing_f) in enumerate(wing_entries):
        row = {"step": step}

        # ── Wing coefficients ──────────────────────────────────────────
        row["wing_aoa"]    = wing_f.get("alpha")
        row["wing_mach"]   = wing_f.get("mach")
        row["wing_CL"]     = wing_f.get("CL")
        row["wing_CD_VFP"] = wing_f.get("CD_VFP")
        row["wing_CD_IBE"] = wing_f.get("CD_IBE")
        row["wing_CM_VFP"] = wing_f.get("CM_VFP")
        row["wing_Cdv"]    = wing_f.get("Cdv")
        row["wing_Cdi"]    = wing_f.get("Cdi")

        # ── Wing wave drag ─────────────────────────────────────────────
        wing_cdw = _extract_wavedrag_cdw(wing_wd_path) if wing_wd_path else None
        row["wing_CDW"] = wing_cdw

        # ── Wing CD total = Cdi + Cdv + CDW_wave ──────────────────────
        row["wing_CD_total"] = (
            (wing_f.get("Cdi") or 0.0)
            + (wing_f.get("Cdv") or 0.0)
            + (wing_cdw or 0.0)
        )

        # ── Tail (if available at this step index) ─────────────────────
        tail_cdw = None
        tail_f   = None
        if step < len(tail_entries):
            tail_fp, tail_wd_path, tail_f = tail_entries[step]

            row["tail_alphat"] = tail_f.get("alpha")
            row["tail_mach"]   = tail_f.get("mach")
            row["tail_CL"]     = tail_f.get("CL")
            row["tail_CD_VFP"] = tail_f.get("CD_VFP")
            row["tail_CD_IBE"] = tail_f.get("CD_IBE")
            row["tail_CM_VFP"] = tail_f.get("CM_VFP")
            row["tail_Cdv"]    = tail_f.get("Cdv")
            row["tail_Cdi"]    = tail_f.get("Cdi")

            tail_cdw = _extract_wavedrag_cdw(tail_wd_path) if tail_wd_path else None
            row["tail_CDW"]      = tail_cdw
            row["tail_CD_total"] = (
                (tail_f.get("Cdi") or 0.0)
                + (tail_f.get("Cdv") or 0.0)
                + (tail_cdw or 0.0)
            )

            # ── Aircraft totals ────────────────────────────────────────
            w_cl = wing_f.get("CL") or 0.0
            t_cl = tail_f.get("CL") or 0.0
            row["total_CL"] = (
                w_cl + t_cl
                if wing_f.get("CL") is not None or tail_f.get("CL") is not None
                else None
            )
            row["total_CM_VFP"] = (
                (wing_f.get("CM_VFP") or 0.0)
                + (tail_f.get("CM_VFP") or 0.0)
            )
            row["total_CD"] = (
                row["wing_CD_total"] + row["tail_CD_total"]
            )
        else:
            # Wing-only total
            row["total_CL"]     = wing_f.get("CL")
            row["total_CM_VFP"] = wing_f.get("CM_VFP")
            row["total_CD"]     = row["wing_CD_total"]

        rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Excel writer  (same style as forces_to_excel.py / vfp-engine.py)
# ---------------------------------------------------------------------------

def export_excel(rows, sim_name, out_path):
    """Write *rows* to a formatted Excel file at *out_path*."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is not installed.  Run:  pip install openpyxl")
        return

    if not rows:
        print("No rows to write – Excel file not created.")
        return

    # ── Colour palette ───────────────────────────────────────────────────
    CLR_HEADER_DARK = "FF1F3864"
    CLR_WING_GRP    = "FF375623"
    CLR_TAIL_GRP    = "FF7030A0"
    CLR_TOTAL_GRP   = "FF833C00"
    CLR_WHITE       = "FFFFFFFF"
    CLR_GREY_ROW    = "FFF2F2F2"

    def _thin_border():
        t = Side(style="thin")
        return Border(left=t, right=t, top=t, bottom=t)

    def _fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def _hdr_font(bold=True, color=CLR_WHITE, size=10):
        return Font(name="Calibri", bold=bold, color=color, size=size)

    def _body_font():
        return Font(name="Calibri", size=10)

    def _center():
        return Alignment(horizontal="center", vertical="center")

    def _right():
        return Alignment(horizontal="right",  vertical="center")

    def _left():
        return Alignment(horizontal="left",   vertical="center")

    def _write_merged(ws, row, col, title, ncols, fill_color):
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + ncols - 1)
        c = ws.cell(row=row, column=col, value=title)
        c.font      = _hdr_font(bold=True, size=10)
        c.fill      = _fill(fill_color)
        c.alignment = _center()
        c.border    = _thin_border()

    def _write_col_hdr(ws, row, col, label, fill_color):
        c = ws.cell(row=row, column=col, value=label)
        c.font      = _hdr_font(color=CLR_WHITE, size=9)
        c.fill      = _fill(fill_color)
        c.alignment = _center()
        c.border    = _thin_border()

    def _write_data(ws, row, col, value, alt=False):
        bg = CLR_GREY_ROW if alt else CLR_WHITE
        c  = ws.cell(row=row, column=col, value=value)
        c.font   = _body_font()
        c.fill   = _fill(bg)
        c.border = _thin_border()
        if isinstance(value, float):
            c.number_format = "0.000000"
            c.alignment     = _right()
        elif isinstance(value, int):
            c.alignment = _right()
        else:
            c.alignment = _left()

    # ── Column definitions ───────────────────────────────────────────────
    # (group_label, group_color, column_header, row_dict_key)
    COLS = [
        (None,    CLR_HEADER_DARK, "Step",          "step"),
        (None,    CLR_WING_GRP,    "Wing AoA (°)",  "wing_aoa"),
        (None,    CLR_WING_GRP,    "Wing Mach",     "wing_mach"),
        ("WING",  CLR_WING_GRP,    "CL",            "wing_CL"),
        ("WING",  CLR_WING_GRP,    "CD (VFP)",      "wing_CD_VFP"),
        ("WING",  CLR_WING_GRP,    "CD (IBE)",      "wing_CD_IBE"),
        ("WING",  CLR_WING_GRP,    "CM (VFP)",      "wing_CM_VFP"),
        ("WING",  CLR_WING_GRP,    "Cd_visc",       "wing_Cdv"),
        ("WING",  CLR_WING_GRP,    "Cd_i (vortex)", "wing_Cdi"),
        ("WING",  CLR_WING_GRP,    "CDW (wave)",    "wing_CDW"),
        ("WING",  CLR_WING_GRP,    "CD_total",      "wing_CD_total"),
    ]

    has_tail = any(r.get("tail_alphat") is not None for r in rows)
    if has_tail:
        COLS += [
            (None,    CLR_TAIL_GRP,   "ε (°)",         "epsilon_deg"),
            (None,    CLR_TAIL_GRP,   "Tail AoA (°)",  "tail_alphat"),
            (None,    CLR_TAIL_GRP,   "Tail Mach",     "tail_mach"),
            ("TAIL",  CLR_TAIL_GRP,   "CL",            "tail_CL"),
            ("TAIL",  CLR_TAIL_GRP,   "CD (VFP)",      "tail_CD_VFP"),
            ("TAIL",  CLR_TAIL_GRP,   "CD (IBE)",      "tail_CD_IBE"),
            ("TAIL",  CLR_TAIL_GRP,   "CM (VFP)",      "tail_CM_VFP"),
            ("TAIL",  CLR_TAIL_GRP,   "Cd_visc",       "tail_Cdv"),
            ("TAIL",  CLR_TAIL_GRP,   "CDW (wave)",    "tail_CDW"),
            ("TAIL",  CLR_TAIL_GRP,   "CD_total",      "tail_CD_total"),
            ("TOTAL", CLR_TOTAL_GRP,  "CL",            "total_CL"),
            ("TOTAL", CLR_TOTAL_GRP,  "CD",            "total_CD"),
            ("TOTAL", CLR_TOTAL_GRP,  "CM (VFP)",      "total_CM_VFP"),
        ]

    ncols = len(COLS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AutoRunner Polar"
    ws.sheet_view.showGridLines = True

    # Row 1 – main title
    ws.row_dimensions[1].height = 24
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=ncols)
    tc = ws.cell(row=1, column=1, value=f"VFP AutoRunner Results – {sim_name}")
    tc.font      = _hdr_font(bold=True, color=CLR_WHITE, size=13)
    tc.fill      = _fill(CLR_HEADER_DARK)
    tc.alignment = _center()
    tc.border    = _thin_border()

    # Row 2 – group headers
    ws.row_dimensions[2].height = 16
    group_runs = []
    run_start = 1
    run_label, run_color = COLS[0][0], COLS[0][1]
    for j, (grp, clr, _, _k) in enumerate(COLS[1:], start=2):
        if grp != run_label or clr != run_color:
            group_runs.append((run_start, j - 1, run_label, run_color))
            run_start = j
            run_label = grp
            run_color = clr
    group_runs.append((run_start, ncols, run_label, run_color))

    for (cs, ce, lbl, clr) in group_runs:
        if lbl:
            _write_merged(ws, 2, cs, lbl, ce - cs + 1, clr)
        else:
            for col in range(cs, ce + 1):
                c = ws.cell(row=2, column=col, value="")
                c.fill   = _fill(clr)
                c.border = _thin_border()

    # Row 3 – column headers
    ws.row_dimensions[3].height = 18
    ws.freeze_panes = "A4"
    for j, (_grp, clr, lbl, _k) in enumerate(COLS, start=1):
        _write_col_hdr(ws, 3, j, lbl, clr)

    ws.auto_filter.ref = f"A3:{get_column_letter(ncols)}3"

    # Data rows
    for idx, row in enumerate(rows):
        r   = 4 + idx
        alt = (idx % 2 == 1)
        ws.row_dimensions[r].height = 15
        for j, (_grp, _clr, _lbl, key) in enumerate(COLS, start=1):
            _write_data(ws, r, j, row.get(key), alt=alt)

    # Column widths
    KEY_WIDTHS = {
        "step": 6, "wing_aoa": 11, "wing_mach": 10,
        "wing_CL": 12, "wing_CD_VFP": 12, "wing_CD_IBE": 12,
        "wing_CM_VFP": 12, "wing_Cdv": 12, "wing_Cdi": 13,
        "wing_CDW": 12, "wing_CD_total": 13,
        "epsilon_deg": 10, "tail_alphat": 12, "tail_mach": 10,
        "tail_CL": 12, "tail_CD_VFP": 12, "tail_CD_IBE": 12,
        "tail_CM_VFP": 12, "tail_Cdv": 12, "tail_CDW": 12,
        "tail_CD_total": 13,
        "total_CL": 12, "total_CD": 12, "total_CM_VFP": 13,
    }
    for j, (_grp, _clr, _lbl, key) in enumerate(COLS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = KEY_WIDTHS.get(key, 12)

    wb.save(out_path)
    print(f"Excel file saved to:\n  {out_path}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(sim_name, project_root):
    sim_base = os.path.join(project_root, "data", "Simulations", sim_name)
    wing_dir = os.path.join(sim_base, "wing")
    tail_dir = os.path.join(sim_base, "tail")

    if not os.path.isdir(wing_dir):
        print(f"ERROR: wing directory not found: {wing_dir}")
        sys.exit(1)

    print(f"Scanning wing directory: {wing_dir}")
    wing_entries = _scan_forces_files(wing_dir)
    if not wing_entries:
        print("ERROR: no readable .forces files found in wing directory.")
        sys.exit(1)
    print(f"  Found {len(wing_entries)} wing step(s)")

    tail_entries = []
    if os.path.isdir(tail_dir):
        print(f"Scanning tail directory: {tail_dir}")
        tail_entries = _scan_forces_files(tail_dir)
        print(f"  Found {len(tail_entries)} tail step(s)")

        if tail_entries and len(tail_entries) != len(wing_entries):
            print(
                f"  WARNING: wing step count ({len(wing_entries)}) != "
                f"tail step count ({len(tail_entries)}). "
                "Tail entries will be matched by index up to the shorter list."
            )
    else:
        print("No tail directory found – wing-only export.")

    rows = build_rows(wing_entries, tail_entries)
    print(f"Built {len(rows)} row(s).")

    out_path = os.path.join(
        project_root, "data", "Simulations",
        f"{sim_name}_AutoRunner_Forces.xlsx"
    )
    export_excel(rows, sim_name, out_path)


def main():
    parser = argparse.ArgumentParser(
        description=(
            "Reconstruct the AutoRunner Excel polar export from an existing "
            "simulation folder."
        )
    )
    parser.add_argument(
        "sim_name",
        help="Simulation name (subfolder of data/Simulations/)."
    )
    parser.add_argument(
        "--project-root", "-r",
        default=None,
        help=(
            "Root of the VFP-Python repository. "
            "Defaults to the parent of the directory containing this script."
        ),
    )
    args = parser.parse_args()

    project_root = args.project_root or os.path.dirname(_SCRIPT_DIR)
    run(args.sim_name, project_root)


if __name__ == "__main__":
    main()
